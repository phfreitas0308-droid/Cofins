# -*- coding: utf-8 -*-
"""
gerar_exemplos.py
=================
Gera dois arquivos .xlsx de EXEMPLO com o layout esperado pelo motor:

  1) balancete.xlsx          -> o balancete contábil
  2) cadastro_cofins.xlsx    -> o cadastro/de-para das contas para PIS/Cofins

Rode este script uma vez para ter arquivos de teste. Depois substitua pelos
seus arquivos reais mantendo os MESMOS nomes de coluna.
"""

from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# --------------------------------------------------------------------------- #
# LAYOUT ESPERADO
#
# balancete.xlsx (aba "Balancete"):
#   Conta | Descricao | Saldo_Anterior | Debito | Credito | Saldo_Atual
#     - Conta: código da conta contábil (texto). Ex.: "3.1.1.001"
#     - Saldo_Atual: saldo final do período (numérico). É o valor usado na apuração.
#       Convenção de sinal: receitas com saldo POSITIVO; despesas com saldo POSITIVO
#       (o motor usa o valor absoluto por classificação, então sinais invertidos
#        são detectados e reportados no log).
#
# cadastro_cofins.xlsx (aba "Cadastro"):
#   Conta | Descricao | Natureza | Classificacao | CST_PIS | CST_COFINS | Gera_Credito | Observacao
#     - Natureza: RECEITA | DESPESA | PATRIMONIAL
#     - Classificacao: RECEITA_TRIBUTADA | RECEITA_EXCLUIDA | DESPESA_DEDUTIVEL |
#                      DESPESA_NAO_DEDUTIVEL | NAO_CLASSIFICADA
#     - Gera_Credito: S | N  (só relevante no regime não-cumulativo)
# --------------------------------------------------------------------------- #

AZUL = "1F4E78"
CINZA = "D9D9D9"


def _cabecalho(ws, colunas):
    ws.append(colunas)
    for c in range(1, len(colunas) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar(ws):
    for col in ws.columns:
        largura = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(largura + 2, 12), 45)


# ---- Balancete de exemplo ------------------------------------------------- #
# (Conta, Descrição, Saldo_Atual)
BALANCETE = [
    # Receitas (grupo 3)
    ("3.1.1.001", "Receita de Venda de Mercadorias", 2_500_000.00),
    ("3.1.1.002", "Receita de Prestação de Serviços", 800_000.00),
    ("3.1.1.003", "Receita de Exportação de Mercadorias", 400_000.00),
    ("3.1.2.001", "Vendas Canceladas", 60_000.00),
    ("3.1.2.002", "Descontos Incondicionais Concedidos", 45_000.00),
    ("3.1.3.001", "IPI sobre Vendas", 90_000.00),
    ("3.1.3.002", "ICMS Substituição Tributária", 30_000.00),
    ("3.2.1.001", "Receita Financeira - Juros Ativos", 25_000.00),
    ("3.2.2.001", "Reversão de Provisões", 15_000.00),
    ("3.2.3.001", "Receita de Venda de Imobilizado", 20_000.00),
    ("3.9.9.001", "Outras Receitas Operacionais", 12_000.00),  # ambígua -> log
    # Custos e despesas (grupo 4)
    ("4.1.1.001", "Compra de Mercadoria para Revenda", 900_000.00),
    ("4.1.1.002", "Matéria-Prima Consumida", 350_000.00),
    ("4.1.2.001", "Energia Elétrica da Produção", 70_000.00),
    ("4.1.2.002", "Aluguel de Galpão Industrial", 120_000.00),
    ("4.1.2.003", "Frete sobre Compras", 40_000.00),
    ("4.1.2.004", "Depreciação de Máquinas", 55_000.00),
    ("4.2.1.001", "Salários e Ordenados", 420_000.00),
    ("4.2.1.002", "Encargos Sociais - FGTS/INSS", 140_000.00),
    ("4.2.2.001", "Despesa Financeira - Juros de Empréstimo", 60_000.00),
    ("4.2.3.001", "Propaganda e Publicidade", 35_000.00),
    ("4.2.3.002", "Material de Escritório", 8_000.00),
    ("4.2.4.001", "Despesas com Viagens", 22_000.00),
    ("4.9.9.001", "Outras Despesas Diversas", 18_000.00),  # ambígua -> log
    # Patrimonial (não entra na apuração)
    ("1.1.1.001", "Caixa e Equivalentes de Caixa", 300_000.00),
    ("2.3.1.001", "Capital Social", 1_000_000.00),
]

# ---- Cadastro Cofins de exemplo ------------------------------------------- #
# Deixamos DE PROPÓSITO algumas contas do balancete FORA do cadastro
# (3.2.3.001, 3.9.9.001, 4.2.4.001, 4.9.9.001) para exercitar a classificação
# automática e o log de inconsistências.
# (Conta, Descrição, Natureza, Classificação, CST_PIS, CST_COFINS, Gera_Credito, Obs)
CADASTRO = [
    ("3.1.1.001", "Receita de Venda de Mercadorias", "RECEITA", "RECEITA_TRIBUTADA", "01", "01", "N", ""),
    ("3.1.1.002", "Receita de Prestação de Serviços", "RECEITA", "RECEITA_TRIBUTADA", "01", "01", "N", ""),
    ("3.1.1.003", "Receita de Exportação de Mercadorias", "RECEITA", "RECEITA_EXCLUIDA", "08", "08", "N", "Exportação - não incidência"),
    ("3.1.2.001", "Vendas Canceladas", "RECEITA", "RECEITA_EXCLUIDA", "01", "01", "N", "Exclusão da base"),
    ("3.1.2.002", "Descontos Incondicionais Concedidos", "RECEITA", "RECEITA_EXCLUIDA", "01", "01", "N", "Exclusão da base"),
    ("3.1.3.001", "IPI sobre Vendas", "RECEITA", "RECEITA_EXCLUIDA", "01", "01", "N", "IPI não integra a base"),
    ("3.1.3.002", "ICMS Substituição Tributária", "RECEITA", "RECEITA_EXCLUIDA", "01", "01", "N", "ICMS-ST exclui base"),
    ("3.2.1.001", "Receita Financeira - Juros Ativos", "RECEITA", "RECEITA_TRIBUTADA", "01", "01", "N", "Receita financeira (não-cumulativo)"),
    ("3.2.2.001", "Reversão de Provisões", "RECEITA", "RECEITA_EXCLUIDA", "08", "08", "N", "Reversão não tributada"),
    # 3.2.3.001 (Venda de Imobilizado)  -> AUSENTE de propósito
    # 3.9.9.001 (Outras Receitas)       -> AUSENTE de propósito
    ("4.1.1.001", "Compra de Mercadoria para Revenda", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Crédito sobre revenda"),
    ("4.1.1.002", "Matéria-Prima Consumida", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Insumo"),
    ("4.1.2.001", "Energia Elétrica da Produção", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Energia gera crédito"),
    ("4.1.2.002", "Aluguel de Galpão Industrial", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Aluguel PJ gera crédito"),
    ("4.1.2.003", "Frete sobre Compras", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Frete de insumo"),
    ("4.1.2.004", "Depreciação de Máquinas", "DESPESA", "DESPESA_DEDUTIVEL", "50", "50", "S", "Encargo de depreciação"),
    ("4.2.1.001", "Salários e Ordenados", "DESPESA", "DESPESA_NAO_DEDUTIVEL", "70", "70", "N", "Mão de obra não gera crédito"),
    ("4.2.1.002", "Encargos Sociais - FGTS/INSS", "DESPESA", "DESPESA_NAO_DEDUTIVEL", "70", "70", "N", ""),
    ("4.2.2.001", "Despesa Financeira - Juros de Empréstimo", "DESPESA", "DESPESA_NAO_DEDUTIVEL", "70", "70", "N", ""),
    ("4.2.3.001", "Propaganda e Publicidade", "DESPESA", "DESPESA_NAO_DEDUTIVEL", "70", "70", "N", ""),
    ("4.2.3.002", "Material de Escritório", "DESPESA", "DESPESA_NAO_DEDUTIVEL", "70", "70", "N", ""),
    # 4.2.4.001 (Viagens)          -> AUSENTE de propósito
    # 4.9.9.001 (Outras Despesas)  -> AUSENTE de propósito
    ("1.1.1.001", "Caixa e Equivalentes de Caixa", "PATRIMONIAL", "NAO_CLASSIFICADA", "", "", "N", "Fora da apuração"),
    ("2.3.1.001", "Capital Social", "PATRIMONIAL", "NAO_CLASSIFICADA", "", "", "N", "Fora da apuração"),
]


def gerar_balancete(caminho: str = "balancete.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balancete"
    _cabecalho(ws, ["Conta", "Descricao", "Saldo_Anterior", "Debito", "Credito", "Saldo_Atual"])
    for conta, desc, saldo in BALANCETE:
        # Saldo_Anterior/Debito/Credito são ilustrativos; a apuração usa Saldo_Atual.
        ws.append([conta, desc, 0.0, 0.0, 0.0, saldo])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=6):
        for cel in row:
            cel.number_format = "#,##0.00"
            cel.font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        for cel in row:
            cel.font = Font(name="Arial", size=10)
    _ajustar(ws)
    ws.freeze_panes = "A2"
    wb.save(caminho)
    return caminho


def gerar_cadastro(caminho: str = "cadastro_cofins.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cadastro"
    _cabecalho(ws, ["Conta", "Descricao", "Natureza", "Classificacao",
                    "CST_PIS", "CST_COFINS", "Gera_Credito", "Observacao"])
    for linha in CADASTRO:
        ws.append(list(linha))
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            cel.font = Font(name="Arial", size=10)
    _ajustar(ws)
    ws.freeze_panes = "A2"
    wb.save(caminho)
    return caminho


if __name__ == "__main__":
    import os, sys
    destino = sys.argv[1] if len(sys.argv) > 1 else "."
    os.makedirs(destino, exist_ok=True)
    b = gerar_balancete(os.path.join(destino, "balancete.xlsx"))
    c = gerar_cadastro(os.path.join(destino, "cadastro_cofins.xlsx"))
    print(f"Gerado: {b}")
    print(f"Gerado: {c}")
