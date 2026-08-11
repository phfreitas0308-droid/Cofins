# -*- coding: utf-8 -*-
"""
relatorio_excel.py
==================
Gera o relatório Excel da apuração com as abas:

  1. Dashboard              -> KPIs + gráficos
  2. Apuração               -> base de cálculo e tributos (com FÓRMULAS vivas)
  3. Balancete Classificado -> cada conta com sua classificação e CST
  4. Log de Inconsistências -> pontos que exigem revisão
  5. Auditoria Tributária   -> trilha da decisão de cada conta e da contribuição

As abas de Apuração e Dashboard usam FÓRMULAS (SUMIFS) referenciando o
Balancete Classificado, de modo que o arquivo recalcula sozinho se você
alterar um saldo ou uma classificação dentro do próprio Excel.
"""

from __future__ import annotations
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

import regras
from regras import (
    RECEITA_TRIBUTADA, RECEITA_EXCLUIDA, DESPESA_DEDUTIVEL,
    DESPESA_NAO_DEDUTIVEL, NAO_CLASSIFICADA, ROTULOS,
)

if TYPE_CHECKING:
    from apuracao_pis_cofins import Apuracao

# ----------------------------- paleta / estilos ---------------------------- #
AZUL_ESCURO = "1F4E78"
AZUL_MEDIO = "2E75B6"
AZUL_CLARO = "DDEBF7"
CINZA_CAB = "44546A"
CINZA_LINHA = "F2F2F2"
VERDE = "548235"
VERDE_CLARO = "E2EFDA"
VERMELHO = "C00000"
VERMELHO_CLARO = "FCE4E4"
AMARELO = "FFF2CC"
LARANJA = "ED7D31"
BRANCO = "FFFFFF"

FONTE = "Arial"
MOEDA = 'R$ #,##0.00;[RED]-R$ #,##0.00'
PCT = "0.00%"

thin = Side(style="thin", color="BFBFBF")
BORDA = Border(left=thin, right=thin, top=thin, bottom=thin)

COR_CLASSIF = {
    RECEITA_TRIBUTADA: VERDE_CLARO,
    RECEITA_EXCLUIDA: AZUL_CLARO,
    DESPESA_DEDUTIVEL: AMARELO,
    DESPESA_NAO_DEDUTIVEL: VERMELHO_CLARO,
    NAO_CLASSIFICADA: CINZA_LINHA,
}
COR_SEVERIDADE = {"Alta": VERMELHO_CLARO, "Média": AMARELO, "Baixa": VERDE_CLARO}


def _titulo(ws, cel, texto, tamanho=14):
    ws[cel] = texto
    ws[cel].font = Font(name=FONTE, size=tamanho, bold=True, color=AZUL_ESCURO)


def _cab(cel, cor=CINZA_CAB):
    cel.font = Font(name=FONTE, size=10, bold=True, color=BRANCO)
    cel.fill = PatternFill("solid", fgColor=cor)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = BORDA


def _cell(ws, ref, valor, *, bold=False, cor=None, fill=None, fmt=None,
          size=10, align=None, borda=False, wrap=False):
    c = ws[ref]
    c.value = valor
    c.font = Font(name=FONTE, size=size, bold=bold, color=cor or "000000")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    elif wrap:
        c.alignment = Alignment(vertical="center", wrap_text=True)
    if borda:
        c.border = BORDA
    return c


def _ajustar_larguras(ws, larguras: dict):
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w


# ======================================================================== #
#  ABA: Balancete Classificado
# ======================================================================== #
def _aba_balancete(wb, apur: "Apuracao"):
    ws = wb.create_sheet("Balancete Classificado")
    headers = ["Conta", "Descrição", "Natureza", "Classificação", "Valor Apuração",
               "CST PIS", "CST COFINS", "Gera Crédito", "Origem", "Confiança", "Chave"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        _cab(c)
    r = 2
    for conta in apur.contas:
        valor = abs(conta.saldo)
        ws.cell(row=r, column=1, value=conta.conta)
        ws.cell(row=r, column=2, value=conta.descricao)
        ws.cell(row=r, column=3, value=conta.natureza)
        ws.cell(row=r, column=4, value=ROTULOS.get(conta.classificacao, conta.classificacao))
        cval = ws.cell(row=r, column=5, value=round(valor, 2))
        cval.number_format = MOEDA
        ws.cell(row=r, column=6, value=conta.cst_pis)
        ws.cell(row=r, column=7, value=conta.cst_cofins)
        ws.cell(row=r, column=8, value="Sim" if conta.gera_credito else "Não")
        ws.cell(row=r, column=9, value=conta.origem)
        ws.cell(row=r, column=10, value=conta.confianca)
        ws.cell(row=r, column=11, value=conta.classificacao)  # chave p/ SUMIFS
        # estilo linha
        fill = COR_CLASSIF.get(conta.classificacao, BRANCO)
        for j in range(1, 12):
            cel = ws.cell(row=r, column=j)
            cel.font = Font(name=FONTE, size=9)
            cel.border = BORDA
            if j in (3, 6, 7, 8, 9, 10):
                cel.alignment = Alignment(horizontal="center")
            if fill and j == 4:
                cel.fill = PatternFill("solid", fgColor=fill)
        r += 1
    n = r - 1
    # total
    ws.cell(row=r, column=4, value="TOTAL").font = Font(name=FONTE, bold=True)
    ctot = ws.cell(row=r, column=5, value=f"=SUM(E2:E{n})" if n >= 2 else 0)
    ctot.number_format = MOEDA
    ctot.font = Font(name=FONTE, bold=True)
    ws.freeze_panes = "A2"
    ws.column_dimensions["K"].hidden = True
    _ajustar_larguras(ws, {"A": 14, "B": 42, "C": 13, "D": 24, "E": 16,
                           "F": 9, "G": 10, "H": 12, "I": 14, "J": 11, "K": 10})
    return ws, n


# ======================================================================== #
#  ABA: Apuração (com fórmulas)
# ======================================================================== #
def _aba_apuracao(wb, apur: "Apuracao", n_contas: int):
    ws = wb.create_sheet("Apuração")
    bc = "'Balancete Classificado'"
    faixa_val = f"{bc}!$E$2:$E${max(n_contas+1,2)}"
    faixa_chave = f"{bc}!$K$2:$K${max(n_contas+1,2)}"
    faixa_cred = f"{bc}!$H$2:$H${max(n_contas+1,2)}"

    _titulo(ws, "A1", "APURAÇÃO DE PIS E COFINS", 16)
    _cell(ws, "A2", f"Empresa: {apur.cfg.empresa or '—'}", size=10)
    _cell(ws, "A3", f"CNPJ: {apur.cfg.cnpj or '—'}    |    Competência: "
                    f"{apur.cfg.competencia or '—'}    |    Regime: "
                    f"{apur.cfg.regime.replace('_',' ')}", size=10, bold=True)

    # ---- Parâmetros (inputs em azul) ----
    _cell(ws, "A5", "PARÂMETROS (editáveis)", bold=True, cor=BRANCO, fill=CINZA_CAB)
    ws.merge_cells("A5:B5")
    aliq = apur.cfg.aliquotas()
    _cell(ws, "A6", "Alíquota PIS")
    _cell(ws, "B6", aliq.pis, cor="0000FF", fmt=PCT, align="right")
    _cell(ws, "A7", "Alíquota COFINS")
    _cell(ws, "B7", aliq.cofins, cor="0000FF", fmt=PCT, align="right")
    _cell(ws, "A8", "Permite crédito (1=sim, 0=não)")
    _cell(ws, "B8", 1 if aliq.permite_credito else 0, cor="0000FF", align="right")

    # ---- Composição da Receita ----
    r = 10
    _cell(ws, f"A{r}", "COMPOSIÇÃO DA RECEITA", bold=True, cor=BRANCO, fill=AZUL_MEDIO)
    ws.merge_cells(f"A{r}:B{r}")
    linhas_receita = [
        ("Receita Tributada", f'=SUMIFS({faixa_val},{faixa_chave},"{RECEITA_TRIBUTADA}")'),
        ("(+) Receita Excluída", f'=SUMIFS({faixa_val},{faixa_chave},"{RECEITA_EXCLUIDA}")'),
        ("(=) Receita Bruta Total", f"=B{r+1}+B{r+2}"),
    ]
    for i, (rot, form) in enumerate(linhas_receita):
        rr = r + 1 + i
        _cell(ws, f"A{rr}", rot, bold=(i == 2))
        _cell(ws, f"B{rr}", form, fmt=MOEDA, align="right", bold=(i == 2),
              fill=AZUL_CLARO if i == 2 else None, borda=True)
    lin_rt, lin_re, lin_rb = r+1, r+2, r+3

    # ---- Despesas ----
    r = lin_rb + 2
    _cell(ws, f"A{r}", "DESPESAS", bold=True, cor=BRANCO, fill=AZUL_MEDIO)
    ws.merge_cells(f"A{r}:B{r}")
    _cell(ws, f"A{r+1}", "Despesa Dedutível (base potencial de crédito)")
    _cell(ws, f"B{r+1}", f'=SUMIFS({faixa_val},{faixa_chave},"{DESPESA_DEDUTIVEL}")',
          fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"A{r+2}", "Despesa Dedutível que GERA crédito")
    _cell(ws, f"B{r+2}",
          f'=SUMIFS({faixa_val},{faixa_chave},"{DESPESA_DEDUTIVEL}",{faixa_cred},"Sim")',
          fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"A{r+3}", "Despesa Não Dedutível")
    _cell(ws, f"B{r+3}", f'=SUMIFS({faixa_val},{faixa_chave},"{DESPESA_NAO_DEDUTIVEL}")',
          fmt=MOEDA, align="right", borda=True)
    lin_dd, lin_dc, lin_dn = r+1, r+2, r+3

    # ---- Base de cálculo ----
    r = lin_dn + 2
    _cell(ws, f"A{r}", "BASE DE CÁLCULO", bold=True, cor=BRANCO, fill=AZUL_ESCURO)
    ws.merge_cells(f"A{r}:B{r}")
    _cell(ws, f"A{r+1}", "Base de Cálculo (Receita Tributada)", bold=True)
    _cell(ws, f"B{r+1}", f"=B{lin_rt}", fmt=MOEDA, align="right", bold=True,
          fill=AMARELO, borda=True)
    lin_base = r+1

    # ---- Tributos ----
    r = lin_base + 2
    _cell(ws, f"A{r}", "APURAÇÃO DOS TRIBUTOS", bold=True, cor=BRANCO, fill=AZUL_ESCURO)
    ws.merge_cells(f"A{r}:D{r}")
    hdr = r+1
    for col, txt in zip("ABCD", ["Tributo", "Débito", "Crédito", "A Pagar"]):
        _cab(ws[f"{col}{hdr}"], cor=CINZA_CAB)
        ws[f"{col}{hdr}"].value = txt
    # PIS
    lp = hdr+1
    _cell(ws, f"A{lp}", "PIS", bold=True, borda=True)
    _cell(ws, f"B{lp}", f"=B{lin_base}*$B$6", fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"C{lp}", f"=B{lin_dc}*$B$6*$B$8", fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"D{lp}", f"=MAX(B{lp}-C{lp},0)", fmt=MOEDA, align="right", bold=True,
          fill=VERDE_CLARO, borda=True)
    # COFINS
    lc = hdr+2
    _cell(ws, f"A{lc}", "COFINS", bold=True, borda=True)
    _cell(ws, f"B{lc}", f"=B{lin_base}*$B$7", fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"C{lc}", f"=B{lin_dc}*$B$7*$B$8", fmt=MOEDA, align="right", borda=True)
    _cell(ws, f"D{lc}", f"=MAX(B{lc}-C{lc},0)", fmt=MOEDA, align="right", bold=True,
          fill=VERDE_CLARO, borda=True)
    # Total
    lt = hdr+3
    _cell(ws, f"A{lt}", "TOTAL", bold=True, fill=AZUL_CLARO, borda=True)
    _cell(ws, f"B{lt}", f"=SUM(B{lp}:B{lc})", fmt=MOEDA, align="right", bold=True, fill=AZUL_CLARO, borda=True)
    _cell(ws, f"C{lt}", f"=SUM(C{lp}:C{lc})", fmt=MOEDA, align="right", bold=True, fill=AZUL_CLARO, borda=True)
    _cell(ws, f"D{lt}", f"=SUM(D{lp}:D{lc})", fmt=MOEDA, align="right", bold=True, fill=LARANJA, borda=True)

    _ajustar_larguras(ws, {"A": 42, "B": 18, "C": 18, "D": 18})

    # devolve referências que o Dashboard vai usar
    return {
        "sheet": "Apuração",
        "rt": f"=Apuração!B{lin_rt}", "re": f"=Apuração!B{lin_re}",
        "rb": f"=Apuração!B{lin_rb}", "dd": f"=Apuração!B{lin_dd}",
        "dn": f"=Apuração!B{lin_dn}", "base": f"=Apuração!B{lin_base}",
        "pis_pagar": f"=Apuração!D{lp}", "cofins_pagar": f"=Apuração!D{lc}",
        "total_pagar": f"=Apuração!D{lt}",
        "pis_deb": f"=Apuração!B{lp}", "cofins_deb": f"=Apuração!B{lc}",
        "pis_cred": f"=Apuração!C{lp}", "cofins_cred": f"=Apuração!C{lc}",
        "_rt": lin_rt, "_re": lin_re, "_dd": lin_dd, "_dn": lin_dn,
        "_lp": lp, "_lc": lc, "_lt": lt,
    }


# ======================================================================== #
#  ABA: Dashboard
# ======================================================================== #
def _kpi(ws, col, row, titulo, ref_formula, cor_fill, cor_txt=BRANCO, fmt=MOEDA):
    """Cria um bloco KPI de 2 colunas x 3 linhas a partir de (col,row)."""
    c0 = get_column_letter(col)
    c1 = get_column_letter(col + 1)
    ws.merge_cells(f"{c0}{row}:{c1}{row}")
    ws.merge_cells(f"{c0}{row+1}:{c1}{row+1}")
    t = ws[f"{c0}{row}"]
    t.value = titulo
    t.font = Font(name=FONTE, size=10, bold=True, color=cor_txt)
    t.fill = PatternFill("solid", fgColor=cor_fill)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    v = ws[f"{c0}{row+1}"]
    v.value = ref_formula
    v.number_format = fmt
    v.font = Font(name=FONTE, size=14, bold=True, color=cor_txt)
    v.fill = PatternFill("solid", fgColor=cor_fill)
    v.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (row, row+1):
        for cc in (c0, c1):
            ws[f"{cc}{rr}"].border = BORDA


def _aba_dashboard(wb, apur: "Apuracao", ref: dict):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    _titulo(ws, "B2", "DASHBOARD — APURAÇÃO PIS/COFINS", 18)
    _cell(ws, "B3", f"{apur.cfg.empresa or 'Empresa'}  |  Competência "
                    f"{apur.cfg.competencia or '—'}  |  Regime "
                    f"{apur.cfg.regime.replace('_',' ')}", size=11, bold=True, cor=CINZA_CAB)

    # KPIs (linha 5-6 e 8-9)
    _kpi(ws, 2, 5, "Receita Bruta Total", ref["rb"], AZUL_MEDIO)
    _kpi(ws, 4, 5, "Receita Excluída", ref["re"], AZUL_MEDIO)
    _kpi(ws, 6, 5, "Base de Cálculo", ref["base"], AZUL_ESCURO)
    _kpi(ws, 8, 5, "Total a Recolher", ref["total_pagar"], LARANJA)

    _kpi(ws, 2, 8, "PIS a Pagar", ref["pis_pagar"], VERDE)
    _kpi(ws, 4, 8, "COFINS a Pagar", ref["cofins_pagar"], VERDE)
    _kpi(ws, 6, 8, "Despesa Dedutível", ref["dd"], CINZA_CAB)
    _kpi(ws, 8, 8, "Nº de Inconsistências",
         len(apur.inconsistencias), VERMELHO if apur.inconsistencias else VERDE, fmt="0")

    # ---- Tabelas auxiliares para gráficos (à direita, colunas N+) ----
    # Composição de resultado
    _cell(ws, "N4", "Categoria", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _cell(ws, "O4", "Valor", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    cats = [
        ("Receita Tributada", f"=Apuração!B{ref['_rt']}"),
        ("Receita Excluída", f"=Apuração!B{ref['_re']}"),
        ("Despesa Dedutível", f"=Apuração!B{ref['_dd']}"),
        ("Despesa Não Dedutível", f"=Apuração!B{ref['_dn']}"),
    ]
    for i, (rot, form) in enumerate(cats):
        _cell(ws, f"N{5+i}", rot, borda=True)
        _cell(ws, f"O{5+i}", form, fmt=MOEDA, align="right", borda=True)

    # Tributos débito/crédito/a pagar
    _cell(ws, "N11", "Tributo", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _cell(ws, "O11", "Débito", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _cell(ws, "P11", "Crédito", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _cell(ws, "Q11", "A Pagar", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _cell(ws, "N12", "PIS", borda=True)
    _cell(ws, "O12", f"=Apuração!B{ref['_lp']}", fmt=MOEDA, borda=True)
    _cell(ws, "P12", f"=Apuração!C{ref['_lp']}", fmt=MOEDA, borda=True)
    _cell(ws, "Q12", f"=Apuração!D{ref['_lp']}", fmt=MOEDA, borda=True)
    _cell(ws, "N13", "COFINS", borda=True)
    _cell(ws, "O13", f"=Apuração!B{ref['_lc']}", fmt=MOEDA, borda=True)
    _cell(ws, "P13", f"=Apuração!C{ref['_lc']}", fmt=MOEDA, borda=True)
    _cell(ws, "Q13", f"=Apuração!D{ref['_lc']}", fmt=MOEDA, borda=True)

    # ---- Gráfico de pizza: composição do resultado ----
    pie = PieChart()
    pie.title = "Composição de Receitas e Despesas"
    labels = Reference(ws, min_col=14, min_row=5, max_row=8)
    data = Reference(ws, min_col=15, min_row=4, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 7.5
    pie.width = 12
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "B12")

    # ---- Gráfico de barras: tributos ----
    bar = BarChart()
    bar.type = "col"
    bar.title = "PIS e COFINS — Débito, Crédito e a Pagar"
    dados = Reference(ws, min_col=15, max_col=17, min_row=11, max_row=13)
    cats_ref = Reference(ws, min_col=14, min_row=12, max_row=13)
    bar.add_data(dados, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.height = 7.5
    bar.width = 14
    bar.y_axis.numFmt = "R$ #,##0"
    bar.y_axis.majorGridlines = None
    ws.add_chart(bar, "B28")

    _ajustar_larguras(ws, {"A": 2, "B": 16, "C": 16, "D": 16, "E": 16,
                           "F": 16, "G": 16, "H": 16, "I": 16,
                           "N": 22, "O": 14, "P": 14, "Q": 14})
    # move dashboard para a primeira posição
    wb.move_sheet("Dashboard", -(len(wb.sheetnames) - 1))
    return ws


# ======================================================================== #
#  ABA: Log de Inconsistências
# ======================================================================== #
def _aba_log(wb, apur: "Apuracao"):
    ws = wb.create_sheet("Log de Inconsistências")
    headers = ["#", "Conta", "Descrição", "Valor", "Tipo", "Detalhe",
               "Severidade", "Ação Sugerida"]
    for j, h in enumerate(headers, start=1):
        _cab(ws.cell(row=1, column=j, value=h))
    if not apur.inconsistencias:
        _cell(ws, "A2", "Nenhuma inconsistência encontrada. ✓",
              bold=True, cor=VERDE, size=11)
    ordem = {"Alta": 0, "Média": 1, "Baixa": 2}
    inc = sorted(apur.inconsistencias, key=lambda x: ordem.get(x.severidade, 3))
    r = 2
    for i, item in enumerate(inc, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=item.conta)
        ws.cell(row=r, column=3, value=item.descricao)
        cv = ws.cell(row=r, column=4, value=round(item.saldo, 2))
        cv.number_format = MOEDA
        ws.cell(row=r, column=5, value=item.tipo)
        ws.cell(row=r, column=6, value=item.detalhe)
        cs = ws.cell(row=r, column=7, value=item.severidade)
        ws.cell(row=r, column=8, value=item.acao_sugerida)
        fill = COR_SEVERIDADE.get(item.severidade, BRANCO)
        for j in range(1, 9):
            cel = ws.cell(row=r, column=j)
            cel.font = Font(name=FONTE, size=9)
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top", wrap_text=(j in (6, 8)))
        cs.fill = PatternFill("solid", fgColor=fill)
        cs.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(r-1,1)}"
    _ajustar_larguras(ws, {"A": 5, "B": 14, "C": 38, "D": 15, "E": 26,
                           "F": 50, "G": 12, "H": 42})
    return ws


# ======================================================================== #
#  ABA: Auditoria Tributária
# ======================================================================== #
def _aba_auditoria(wb, apur: "Apuracao"):
    ws = wb.create_sheet("Auditoria Tributária")
    _titulo(ws, "A1", "AUDITORIA TRIBUTÁRIA — TRILHA DE DECISÃO", 14)
    _cell(ws, "A2", "Rastreabilidade de cada conta: como foi classificada, com que "
                    "confiança e sua contribuição para os tributos.", size=9, cor=CINZA_CAB)
    headers = ["Conta", "Descrição", "Classificação", "Origem", "Confiança",
               "Regra Aplicada", "Valor", "CST PIS", "CST COFINS",
               "Base de Crédito", "PIS (linha)", "COFINS (linha)"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        _cab(ws.cell(row=hr, column=j, value=h))
    aliq = apur.cfg.aliquotas()
    r = hr + 1
    for c in apur.contas:
        valor = abs(c.saldo)
        # contribuição de débito da conta (se receita tributada)
        pis_lin = valor * aliq.pis if c.classificacao == RECEITA_TRIBUTADA else 0.0
        cof_lin = valor * aliq.cofins if c.classificacao == RECEITA_TRIBUTADA else 0.0
        # base de crédito (se despesa dedutível que gera crédito e regime permite)
        base_cred = valor if (c.classificacao == DESPESA_DEDUTIVEL and c.gera_credito
                              and aliq.permite_credito) else 0.0
        # crédito reduz o tributo -> mostra como negativo nas colunas de PIS/COFINS
        if base_cred:
            pis_lin = -base_cred * aliq.pis
            cof_lin = -base_cred * aliq.cofins
        vals = [c.conta, c.descricao, ROTULOS.get(c.classificacao, c.classificacao),
                c.origem, c.confianca, c.regra, round(valor, 2), c.cst_pis,
                c.cst_cofins, round(base_cred, 2), round(pis_lin, 2), round(cof_lin, 2)]
        for j, v in enumerate(vals, start=1):
            cel = ws.cell(row=r, column=j, value=v)
            cel.font = Font(name=FONTE, size=9)
            cel.border = BORDA
            cel.alignment = Alignment(vertical="top", wrap_text=(j == 6))
            if j in (7, 10, 11, 12):
                cel.number_format = MOEDA
            if j in (4, 5, 8, 9):
                cel.alignment = Alignment(horizontal="center", vertical="top")
            if c.confianca == "baixa" and j == 5:
                cel.fill = PatternFill("solid", fgColor=VERMELHO_CLARO)
            elif c.confianca == "media" and j == 5:
                cel.fill = PatternFill("solid", fgColor=AMARELO)
        r += 1
    n = r - 1
    # linha de totais
    _cell(ws, f"F{r}", "TOTAIS", bold=True)
    for col in ("G", "J", "K", "L"):
        cc = ws[f"{col}{r}"]
        cc.value = f"=SUM({col}{hr+1}:{col}{n})" if n > hr else 0
        cc.number_format = MOEDA
        cc.font = Font(name=FONTE, bold=True)
        cc.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hr}:L{n}"
    _ajustar_larguras(ws, {"A": 14, "B": 38, "C": 24, "D": 14, "E": 11,
                           "F": 46, "G": 15, "H": 9, "I": 10, "J": 15,
                           "K": 14, "L": 14})
    return ws


# ======================================================================== #
#  Função pública
# ======================================================================== #
def gerar(apur: "Apuracao", caminho: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove a aba default vazia

    _, n_contas = _aba_balancete(wb, apur)
    ref = _aba_apuracao(wb, apur, n_contas)
    _aba_log(wb, apur)
    _aba_auditoria(wb, apur)
    _aba_dashboard(wb, apur, ref)  # cria e move para a frente

    # ordem final: Dashboard, Apuração, Balancete, Log, Auditoria
    ordem = ["Dashboard", "Apuração", "Balancete Classificado",
             "Log de Inconsistências", "Auditoria Tributária"]
    wb._sheets.sort(key=lambda s: ordem.index(s.title) if s.title in ordem else 99)

    wb.save(caminho)
    return caminho
