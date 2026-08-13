# -*- coding: utf-8 -*-
"""
relatorio_if.py
===============
Relatório Excel da apuração de PIS/COFINS de instituição financeira (COSIF),
com as abas: Dashboard, Apuração, Balancete Classificado, Log de
Inconsistências e Auditoria Tributária.

A aba Apuração usa fórmulas SUMIFS ligadas ao Balancete Classificado, então o
arquivo recalcula sozinho se você mudar um saldo ou uma classificação.
"""

from __future__ import annotations
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

from regras_if import (
    RENDA_TRIBUTADA, RENDA_EXCLUIDA, DEDUCAO_INTERMEDIACAO,
    DESPESA_OPERACIONAL, FORA_APURACAO, ROTULOS,
    ALIQUOTA_PIS, ALIQUOTA_COFINS,
)

if TYPE_CHECKING:
    from apuracao_if import Apuracao

AZUL_ESCURO = "1F4E78"; AZUL_MEDIO = "2E75B6"; AZUL_CLARO = "DDEBF7"
CINZA_CAB = "44546A"; VERDE_CLARO = "E2EFDA"; VERMELHO_CLARO = "FCE4E4"
AMARELO = "FFF2CC"; LARANJA = "ED7D31"; VERDE = "548235"; VERMELHO = "C00000"
BRANCO = "FFFFFF"; CINZA_LINHA = "F2F2F2"
FONTE = "Arial"; MOEDA = 'R$ #,##0.00;[RED]-R$ #,##0.00'; PCT = "0.00%"

thin = Side(style="thin", color="BFBFBF")
BORDA = Border(left=thin, right=thin, top=thin, bottom=thin)

COR_CLASSIF = {
    RENDA_TRIBUTADA: VERDE_CLARO,
    RENDA_EXCLUIDA: AZUL_CLARO,
    DEDUCAO_INTERMEDIACAO: AMARELO,
    DESPESA_OPERACIONAL: VERMELHO_CLARO,
    FORA_APURACAO: CINZA_LINHA,
}
COR_SEV = {"Alta": VERMELHO_CLARO, "Média": AMARELO, "Baixa": VERDE_CLARO}


def _titulo(ws, cel, txt, tam=14):
    ws[cel] = txt
    ws[cel].font = Font(name=FONTE, size=tam, bold=True, color=AZUL_ESCURO)


def _cab(cel, cor=CINZA_CAB):
    cel.font = Font(name=FONTE, size=10, bold=True, color=BRANCO)
    cel.fill = PatternFill("solid", fgColor=cor)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = BORDA


def _c(ws, ref, val, *, bold=False, cor=None, fill=None, fmt=None, size=10,
       align=None, borda=False, wrap=False):
    cel = ws[ref]; cel.value = val
    cel.font = Font(name=FONTE, size=size, bold=bold, color=cor or "000000")
    if fill: cel.fill = PatternFill("solid", fgColor=fill)
    if fmt: cel.number_format = fmt
    if align: cel.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    elif wrap: cel.alignment = Alignment(vertical="center", wrap_text=True)
    if borda: cel.border = BORDA
    return cel


def _larg(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w


# ---------------- Balancete Classificado ---------------- #
def _aba_balancete(wb, apur):
    ws = wb.create_sheet("Balancete Classificado")
    heads = ["Conta", "Descrição", "Natureza", "Classificação", "Valor",
             "CST PIS", "CST COFINS", "Origem", "Confiança", "Chave"]
    for j, h in enumerate(heads, 1):
        _cab(ws.cell(row=1, column=j, value=h))
    r = 2
    for c in apur.contas:
        valor = abs(c.saldo)
        ws.cell(row=r, column=1, value=c.conta)
        ws.cell(row=r, column=2, value=c.descricao)
        ws.cell(row=r, column=3, value=c.natureza)
        ws.cell(row=r, column=4, value=ROTULOS.get(c.classificacao, c.classificacao))
        cv = ws.cell(row=r, column=5, value=round(valor, 2)); cv.number_format = MOEDA
        ws.cell(row=r, column=6, value=c.cst_pis)
        ws.cell(row=r, column=7, value=c.cst_cofins)
        ws.cell(row=r, column=8, value=c.origem)
        ws.cell(row=r, column=9, value=c.confianca)
        ws.cell(row=r, column=10, value=c.classificacao)
        fill = COR_CLASSIF.get(c.classificacao, BRANCO)
        for j in range(1, 11):
            cel = ws.cell(row=r, column=j)
            cel.font = Font(name=FONTE, size=9); cel.border = BORDA
            if j in (3, 6, 7, 8, 9): cel.alignment = Alignment(horizontal="center")
            if fill and j == 4: cel.fill = PatternFill("solid", fgColor=fill)
        r += 1
    n = r - 1
    ws.cell(row=r, column=4, value="TOTAL").font = Font(name=FONTE, bold=True)
    ct = ws.cell(row=r, column=5, value=f"=SUM(E2:E{n})" if n >= 2 else 0)
    ct.number_format = MOEDA; ct.font = Font(name=FONTE, bold=True)
    ws.freeze_panes = "A2"; ws.column_dimensions["J"].hidden = True
    _larg(ws, {"A": 15, "B": 48, "C": 13, "D": 34, "E": 16, "F": 9, "G": 10,
               "H": 14, "I": 11, "J": 10})
    return n


# ---------------- Apuração (fórmulas) ---------------- #
def _aba_apuracao(wb, apur, n):
    ws = wb.create_sheet("Apuração")
    bc = "'Balancete Classificado'"
    val = f"{bc}!$E$2:$E${max(n+1,2)}"
    chave = f"{bc}!$J$2:$J${max(n+1,2)}"

    _titulo(ws, "A1", "APURAÇÃO DE PIS E COFINS — INSTITUIÇÃO FINANCEIRA", 15)
    _c(ws, "A2", f"Empresa: {apur.cfg.empresa or '—'}    |    CNPJ: {apur.cfg.cnpj or '—'}", size=10)
    _c(ws, "A3", f"Competência: {apur.cfg.competencia or '—'}    |    Regime: "
       f"cumulativo (instituição financeira)    |    Tipo: {apur.cfg.tipo_instituicao}",
       size=10, bold=True)

    _c(ws, "A5", "PARÂMETROS", bold=True, cor=BRANCO, fill=CINZA_CAB); ws.merge_cells("A5:B5")
    _c(ws, "A6", "Alíquota PIS"); _c(ws, "B6", ALIQUOTA_PIS, cor="0000FF", fmt=PCT, align="right")
    _c(ws, "A7", "Alíquota COFINS"); _c(ws, "B7", ALIQUOTA_COFINS, cor="0000FF", fmt=PCT, align="right")

    r = 9
    _c(ws, f"A{r}", "COMPOSIÇÃO DA RECEITA (grupo 7 - COSIF)", bold=True, cor=BRANCO, fill=AZUL_MEDIO)
    ws.merge_cells(f"A{r}:B{r}")
    _c(ws, f"A{r+1}", "Rendas Tributadas")
    _c(ws, f"B{r+1}", f'=SUMIFS({val},{chave},"{RENDA_TRIBUTADA}")', fmt=MOEDA, align="right", borda=True)
    _c(ws, f"A{r+2}", "(+) Rendas Excluídas / Não Tributadas")
    _c(ws, f"B{r+2}", f'=SUMIFS({val},{chave},"{RENDA_EXCLUIDA}")', fmt=MOEDA, align="right", borda=True)
    _c(ws, f"A{r+3}", "(=) Receita Bruta Operacional", bold=True)
    _c(ws, f"B{r+3}", f"=B{r+1}+B{r+2}", fmt=MOEDA, align="right", bold=True, fill=AZUL_CLARO, borda=True)
    lin_trib, lin_excl, lin_rbo = r+1, r+2, r+3

    r = lin_rbo + 2
    _c(ws, f"A{r}", "DEDUÇÕES DA BASE (art. 3º §6º I - Lei 9.718/98)", bold=True, cor=BRANCO, fill=AZUL_MEDIO)
    ws.merge_cells(f"A{r}:B{r}")
    _c(ws, f"A{r+1}", "(-) Deduções de Intermediação Financeira")
    _c(ws, f"B{r+1}", f'=SUMIFS({val},{chave},"{DEDUCAO_INTERMEDIACAO}")', fmt=MOEDA, align="right", borda=True)
    lin_ded = r+1

    r = lin_ded + 2
    _c(ws, f"A{r}", "BASE DE CÁLCULO", bold=True, cor=BRANCO, fill=AZUL_ESCURO)
    ws.merge_cells(f"A{r}:B{r}")
    _c(ws, f"A{r+1}", "Base de Cálculo (Rendas Tributadas − Deduções)", bold=True)
    _c(ws, f"B{r+1}", f"=MAX(B{lin_trib}-B{lin_ded},0)", fmt=MOEDA, align="right", bold=True, fill=AMARELO, borda=True)
    lin_base = r+1

    r = lin_base + 2
    _c(ws, f"A{r}", "APURAÇÃO DOS TRIBUTOS", bold=True, cor=BRANCO, fill=AZUL_ESCURO)
    ws.merge_cells(f"A{r}:B{r}")
    _c(ws, f"A{r+1}", "PIS (0,65%)", bold=True, borda=True)
    _c(ws, f"B{r+1}", f"=B{lin_base}*$B$6", fmt=MOEDA, align="right", bold=True, fill=VERDE_CLARO, borda=True)
    _c(ws, f"A{r+2}", "COFINS (4,00%)", bold=True, borda=True)
    _c(ws, f"B{r+2}", f"=B{lin_base}*$B$7", fmt=MOEDA, align="right", bold=True, fill=VERDE_CLARO, borda=True)
    _c(ws, f"A{r+3}", "TOTAL A RECOLHER", bold=True, fill=AZUL_CLARO, borda=True)
    _c(ws, f"B{r+3}", f"=B{r+1}+B{r+2}", fmt=MOEDA, align="right", bold=True, fill=LARANJA, borda=True)
    lin_pis, lin_cof, lin_tot = r+1, r+2, r+3

    r = lin_tot + 2
    _c(ws, f"A{r}", "MEMÓRIA (informativo — não afeta a base)", bold=True, cor=BRANCO, fill=CINZA_CAB)
    ws.merge_cells(f"A{r}:B{r}")
    _c(ws, f"A{r+1}", "Despesas Operacionais (pessoal, adm., PCLD, etc.)")
    _c(ws, f"B{r+1}", f'=SUMIFS({val},{chave},"{DESPESA_OPERACIONAL}")', fmt=MOEDA, align="right", borda=True)
    lin_desp = r+1

    r = lin_desp + 2
    _c(ws, f"A{r}", "BASE LEGAL", bold=True, cor=BRANCO, fill=CINZA_CAB); ws.merge_cells(f"A{r}:B{r}")
    txts = [
        "• Regime cumulativo obrigatório: art. 8º, I, da Lei 10.637/2002 e art. 10, I, da",
        "   Lei 10.833/2003 (PJ do §6º do art. 3º da Lei 9.718/98) — sem créditos.",
        "• COFINS 4%: art. 18 da Lei 10.684/2003.   PIS 0,65%.",
        "• Deduções da base: art. 3º, §6º, I, 'a' a 'e', da Lei 9.718/98 (intermediação",
        "   financeira; obrigações por empréstimos/repasses; deságio; perdas com títulos",
        "   exceto ações; perdas em hedge).",
    ]
    for i, t in enumerate(txts, 1):
        _c(ws, f"A{r+i}", t, size=9, cor=CINZA_CAB); ws.merge_cells(f"A{r+i}:B{r+i}")

    _larg(ws, {"A": 52, "B": 20})
    return {"trib": lin_trib, "excl": lin_excl, "rbo": lin_rbo, "ded": lin_ded,
            "base": lin_base, "pis": lin_pis, "cof": lin_cof, "tot": lin_tot, "desp": lin_desp}


# ---------------- Dashboard ---------------- #
def _kpi(ws, col, row, titulo, ref, cor, fmt=MOEDA):
    c0, c1 = get_column_letter(col), get_column_letter(col + 1)
    ws.merge_cells(f"{c0}{row}:{c1}{row}"); ws.merge_cells(f"{c0}{row+1}:{c1}{row+1}")
    t = ws[f"{c0}{row}"]; t.value = titulo
    t.font = Font(name=FONTE, size=10, bold=True, color=BRANCO)
    t.fill = PatternFill("solid", fgColor=cor)
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    v = ws[f"{c0}{row+1}"]; v.value = ref; v.number_format = fmt
    v.font = Font(name=FONTE, size=14, bold=True, color=BRANCO)
    v.fill = PatternFill("solid", fgColor=cor)
    v.alignment = Alignment(horizontal="center", vertical="center")
    for rr in (row, row+1):
        for cc in (c0, c1): ws[f"{cc}{rr}"].border = BORDA


def _aba_dashboard(wb, apur, ref):
    ws = wb.create_sheet("Dashboard"); ws.sheet_view.showGridLines = False
    _titulo(ws, "B2", "DASHBOARD — PIS/COFINS INSTITUIÇÃO FINANCEIRA", 16)
    _c(ws, "B3", f"{apur.cfg.empresa or 'Instituição'}  |  Competência "
       f"{apur.cfg.competencia or '—'}  |  Regime cumulativo", size=11, bold=True, cor=CINZA_CAB)

    A = "Apuração"
    _kpi(ws, 2, 5, "Receita Bruta Operacional", f"={A}!B{ref['rbo']}", AZUL_MEDIO)
    _kpi(ws, 4, 5, "Deduções §6º I", f"={A}!B{ref['ded']}", AZUL_MEDIO)
    _kpi(ws, 6, 5, "Base de Cálculo", f"={A}!B{ref['base']}", AZUL_ESCURO)
    _kpi(ws, 8, 5, "Total a Recolher", f"={A}!B{ref['tot']}", LARANJA)
    _kpi(ws, 2, 8, "PIS (0,65%)", f"={A}!B{ref['pis']}", VERDE)
    _kpi(ws, 4, 8, "COFINS (4,00%)", f"={A}!B{ref['cof']}", VERDE)
    _kpi(ws, 6, 8, "Rendas Excluídas", f"={A}!B{ref['excl']}", CINZA_CAB)
    _kpi(ws, 8, 8, "Nº de Inconsistências", len(apur.inconsistencias),
         VERMELHO if apur.inconsistencias else VERDE, fmt="0")

    # tabela para gráfico de composição
    _c(ws, "N4", "Categoria", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _c(ws, "O4", "Valor", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    cats = [
        ("Rendas Tributadas", f"={A}!B{ref['trib']}"),
        ("Rendas Excluídas", f"={A}!B{ref['excl']}"),
        ("Deduções §6º I", f"={A}!B{ref['ded']}"),
        ("Despesas Operacionais", f"={A}!B{ref['desp']}"),
    ]
    for i, (rot, f) in enumerate(cats):
        _c(ws, f"N{5+i}", rot, borda=True); _c(ws, f"O{5+i}", f, fmt=MOEDA, align="right", borda=True)

    # tabela para gráfico de tributos
    _c(ws, "N11", "Tributo", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _c(ws, "O11", "Valor", bold=True, fill=CINZA_CAB, cor=BRANCO, borda=True)
    _c(ws, "N12", "PIS", borda=True); _c(ws, "O12", f"={A}!B{ref['pis']}", fmt=MOEDA, borda=True)
    _c(ws, "N13", "COFINS", borda=True); _c(ws, "O13", f"={A}!B{ref['cof']}", fmt=MOEDA, borda=True)

    pie = PieChart(); pie.title = "Composição (Rendas e Deduções)"
    labels = Reference(ws, min_col=14, min_row=5, max_row=8)
    data = Reference(ws, min_col=15, min_row=4, max_row=8)
    pie.add_data(data, titles_from_data=True); pie.set_categories(labels)
    pie.height = 7.5; pie.width = 12
    pie.dataLabels = DataLabelList(); pie.dataLabels.showPercent = True
    ws.add_chart(pie, "B12")

    bar = BarChart(); bar.type = "col"; bar.title = "PIS e COFINS a Recolher"
    dados = Reference(ws, min_col=15, min_row=11, max_row=13)
    catref = Reference(ws, min_col=14, min_row=12, max_row=13)
    bar.add_data(dados, titles_from_data=True); bar.set_categories(catref)
    bar.height = 7.5; bar.width = 12; bar.y_axis.numFmt = "R$ #,##0"; bar.legend = None
    ws.add_chart(bar, "B28")

    _larg(ws, {"A": 2, "B": 17, "C": 17, "D": 17, "E": 17, "F": 17, "G": 17,
               "H": 17, "I": 17, "N": 22, "O": 15})
    wb.move_sheet("Dashboard", -(len(wb.sheetnames) - 1))


# ---------------- Log ---------------- #
def _aba_log(wb, apur):
    ws = wb.create_sheet("Log de Inconsistências")
    heads = ["#", "Conta", "Descrição", "Valor", "Tipo", "Detalhe", "Severidade", "Ação Sugerida"]
    for j, h in enumerate(heads, 1): _cab(ws.cell(row=1, column=j, value=h))
    if not apur.inconsistencias:
        _c(ws, "A2", "Nenhuma inconsistência encontrada. ✓", bold=True, cor=VERDE, size=11)
    ordem = {"Alta": 0, "Média": 1, "Baixa": 2}
    inc = sorted(apur.inconsistencias, key=lambda x: ordem.get(x.severidade, 3))
    r = 2
    for i, it in enumerate(inc, 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=it.conta)
        ws.cell(row=r, column=3, value=it.descricao)
        cv = ws.cell(row=r, column=4, value=round(it.saldo, 2)); cv.number_format = MOEDA
        ws.cell(row=r, column=5, value=it.tipo)
        ws.cell(row=r, column=6, value=it.detalhe)
        cs = ws.cell(row=r, column=7, value=it.severidade)
        ws.cell(row=r, column=8, value=it.acao_sugerida)
        for j in range(1, 9):
            cel = ws.cell(row=r, column=j); cel.font = Font(name=FONTE, size=9)
            cel.border = BORDA; cel.alignment = Alignment(vertical="top", wrap_text=(j in (6, 8)))
        cs.fill = PatternFill("solid", fgColor=COR_SEV.get(it.severidade, BRANCO))
        cs.alignment = Alignment(horizontal="center", vertical="center")
        r += 1
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:H{max(r-1,1)}"
    _larg(ws, {"A": 5, "B": 15, "C": 44, "D": 15, "E": 26, "F": 52, "G": 12, "H": 40})


# ---------------- Auditoria ---------------- #
def _aba_auditoria(wb, apur):
    ws = wb.create_sheet("Auditoria Tributária")
    _titulo(ws, "A1", "AUDITORIA TRIBUTÁRIA — TRILHA DE DECISÃO", 14)
    _c(ws, "A2", "Como cada conta foi classificada e sua contribuição para a apuração.",
       size=9, cor=CINZA_CAB)
    heads = ["Conta", "Descrição", "Classificação", "Origem", "Confiança", "Regra Aplicada",
             "Valor", "Compõe base?", "Reduz base?", "PIS", "COFINS"]
    hr = 4
    for j, h in enumerate(heads, 1): _cab(ws.cell(row=hr, column=j, value=h))
    r = hr + 1
    for c in apur.contas:
        valor = abs(c.saldo)
        compoe = valor if c.classificacao == RENDA_TRIBUTADA else 0.0
        reduz = valor if c.classificacao == DEDUCAO_INTERMEDIACAO else 0.0
        pis = compoe * ALIQUOTA_PIS - reduz * ALIQUOTA_PIS
        cof = compoe * ALIQUOTA_COFINS - reduz * ALIQUOTA_COFINS
        vals = [c.conta, c.descricao, ROTULOS.get(c.classificacao, c.classificacao),
                c.origem, c.confianca, c.regra, round(valor, 2),
                "Sim" if compoe else "—", "Sim" if reduz else "—",
                round(pis, 2), round(cof, 2)]
        for j, v in enumerate(vals, 1):
            cel = ws.cell(row=r, column=j, value=v); cel.font = Font(name=FONTE, size=9)
            cel.border = BORDA; cel.alignment = Alignment(vertical="top", wrap_text=(j == 6))
            if j in (7, 10, 11): cel.number_format = MOEDA
            if j in (4, 5, 8, 9): cel.alignment = Alignment(horizontal="center", vertical="top")
            if c.confianca == "baixa" and j == 5: cel.fill = PatternFill("solid", fgColor=VERMELHO_CLARO)
            elif c.confianca == "media" and j == 5: cel.fill = PatternFill("solid", fgColor=AMARELO)
        r += 1
    n = r - 1
    _c(ws, f"F{r}", "TOTAIS", bold=True)
    for col in ("G", "J", "K"):
        cc = ws[f"{col}{r}"]; cc.value = f"=SUM({col}{hr+1}:{col}{n})" if n > hr else 0
        cc.number_format = MOEDA; cc.font = Font(name=FONTE, bold=True)
        cc.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A{hr}:K{n}"
    _larg(ws, {"A": 15, "B": 46, "C": 34, "D": 14, "E": 11, "F": 44, "G": 16,
               "H": 13, "I": 12, "J": 15, "K": 15})


# ---------------- Público ---------------- #
def gerar(apur, caminho: str) -> str:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    n = _aba_balancete(wb, apur)
    ref = _aba_apuracao(wb, apur, n)
    _aba_log(wb, apur)
    _aba_auditoria(wb, apur)
    _aba_dashboard(wb, apur, ref)

    ordem = ["Dashboard", "Apuração", "Balancete Classificado",
             "Log de Inconsistências", "Auditoria Tributária"]
    wb._sheets.sort(key=lambda s: ordem.index(s.title) if s.title in ordem else 99)

    largas = {"Balancete Classificado", "Log de Inconsistências", "Auditoria Tributária"}
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape" if ws.title in largas else "portrait"
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.print_options.horizontalCentered = True
        ws.page_margins = openpyxl.worksheet.page.PageMargins(
            left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    wb.save(caminho)
    return caminho
