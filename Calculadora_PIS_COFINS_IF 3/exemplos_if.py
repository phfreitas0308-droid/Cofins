# -*- coding: utf-8 -*-
"""
exemplos_if.py
==============
Gera arquivos de exemplo no plano COSIF para uma instituição financeira (banco):
  - balancete.xlsx  (grupos 7 = rendas, 8 = despesas, 1/6 = patrimonial)
  - cadastro.xlsx   (classificação PIS/COFINS de cada conta)

LAYOUT ESPERADO
  balancete.xlsx (aba "Balancete"):
     Conta | Descricao | Saldo_Anterior | Debito | Credito | Saldo_Atual
     - a apuração usa a coluna Saldo_Atual.
  cadastro.xlsx (aba "Cadastro"):
     Conta | Descricao | Natureza | Classificacao | CST_PIS | CST_COFINS | Observacao
     - Natureza: RENDA | DESPESA | PATRIMONIAL
     - Classificacao: RENDA_TRIBUTADA | RENDA_EXCLUIDA | DEDUCAO_INTERMEDIACAO |
                      DESPESA_OPERACIONAL | FORA_APURACAO
"""

from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

AZUL = "1F4E78"


def _cab(ws, cols):
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZUL)
        cel.alignment = Alignment(horizontal="center", vertical="center")


def _ajustar(ws):
    for col in ws.columns:
        larg = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(larg + 2, 12), 55)


# (Conta COSIF, Descrição, Saldo_Atual)
BALANCETE = [
    # ---- Grupo 7: RENDAS (receitas) ----
    # Rendas tributadas (compõem a base)
    ("7.1.1.10.00", "Rendas de Operações de Crédito", 8_000_000.00),
    ("7.1.5.00.00", "Rendas de Operações de Arrendamento Mercantil", 500_000.00),
    ("7.1.7.00.00", "Rendas de Títulos e Valores Mobiliários", 2_000_000.00),
    ("7.1.9.00.00", "Rendas de Aplicações Interfinanceiras de Liquidez", 1_200_000.00),
    ("7.1.3.00.00", "Rendas de Operações de Câmbio", 300_000.00),          # ausente no cadastro
    ("7.1.7.50.00", "Rendas de Prestação de Serviços (tarifas)", 1_500_000.00),
    # Rendas excluídas / não tributadas
    ("7.1.9.90.00", "Reversão de Provisões Operacionais", 400_000.00),
    ("7.9.7.00.00", "Resultado de Participações em Coligadas (equiv. patrim.)", 250_000.00),
    ("7.9.9.10.00", "Receita de Prestação de Serviços ao Exterior", 150_000.00),
    ("7.9.9.20.00", "Recuperação de Créditos Baixados como Prejuízo", 200_000.00),
    # ---- Grupo 8: DESPESAS ----
    # Deduções da base - art. 3º, §6º, I, "a" a "e", da Lei 9.718/98
    ("8.1.1.00.00", "Despesas de Operações de Captação no Mercado", 5_000_000.00),   # a
    ("8.1.3.00.00", "Despesas de Obrigações por Empréstimos e Repasses", 1_500_000.00),  # b
    ("8.1.7.10.00", "Deságio na Colocação de Títulos", 200_000.00),                  # c
    ("8.1.7.20.00", "Perdas com Títulos de Renda Fixa (exceto ações)", 300_000.00),  # d
    ("8.1.7.30.00", "Perdas com Instrumentos Financeiros Derivativos (hedge)", 500_000.00),  # e
    # Despesas operacionais (NÃO reduzem a base - apenas informativas)
    ("8.1.7.00.00", "Despesas de Pessoal", 2_500_000.00),
    ("8.1.7.05.00", "Despesas Administrativas", 1_800_000.00),
    ("8.1.7.06.00", "Despesas Tributárias", 400_000.00),
    ("8.1.9.00.00", "Provisão para Créditos de Liquidação Duvidosa (PCLD)", 1_000_000.00),  # ausente no cadastro
    # ---- Patrimonial (fora da apuração) ----
    ("1.1.1.00.00", "Disponibilidades", 3_000_000.00),
    ("6.1.1.00.00", "Capital Social", 10_000_000.00),
]

# Cadastro. Deixamos 7.1.3.00.00 (câmbio) e 8.1.9.00.00 (PCLD) FORA de propósito,
# para exercitar a classificação automática e o log.
CADASTRO = [
    ("7.1.1.10.00", "Rendas de Operações de Crédito", "RENDA", "RENDA_TRIBUTADA", "01", "01", ""),
    ("7.1.5.00.00", "Rendas de Operações de Arrendamento Mercantil", "RENDA", "RENDA_TRIBUTADA", "01", "01", ""),
    ("7.1.7.00.00", "Rendas de Títulos e Valores Mobiliários", "RENDA", "RENDA_TRIBUTADA", "01", "01", ""),
    ("7.1.9.00.00", "Rendas de Aplicações Interfinanceiras de Liquidez", "RENDA", "RENDA_TRIBUTADA", "01", "01", ""),
    ("7.1.7.50.00", "Rendas de Prestação de Serviços (tarifas)", "RENDA", "RENDA_TRIBUTADA", "01", "01", ""),
    ("7.1.9.90.00", "Reversão de Provisões Operacionais", "RENDA", "RENDA_EXCLUIDA", "08", "08", "Reversão não tributada"),
    ("7.9.7.00.00", "Resultado de Participações em Coligadas (equiv. patrim.)", "RENDA", "RENDA_EXCLUIDA", "08", "08", "Equivalência patrimonial"),
    ("7.9.9.10.00", "Receita de Prestação de Serviços ao Exterior", "RENDA", "RENDA_EXCLUIDA", "08", "08", "Exportação de serviços"),
    ("7.9.9.20.00", "Recuperação de Créditos Baixados como Prejuízo", "RENDA", "RENDA_EXCLUIDA", "08", "08", "Recuperação - exclusão"),
    ("8.1.1.00.00", "Despesas de Operações de Captação no Mercado", "DESPESA", "DEDUCAO_INTERMEDIACAO", "", "", "Lei 9.718/98 §6 I 'a'"),
    ("8.1.3.00.00", "Despesas de Obrigações por Empréstimos e Repasses", "DESPESA", "DEDUCAO_INTERMEDIACAO", "", "", "Lei 9.718/98 §6 I 'b'"),
    ("8.1.7.10.00", "Deságio na Colocação de Títulos", "DESPESA", "DEDUCAO_INTERMEDIACAO", "", "", "Lei 9.718/98 §6 I 'c'"),
    ("8.1.7.20.00", "Perdas com Títulos de Renda Fixa (exceto ações)", "DESPESA", "DEDUCAO_INTERMEDIACAO", "", "", "Lei 9.718/98 §6 I 'd'"),
    ("8.1.7.30.00", "Perdas com Instrumentos Financeiros Derivativos (hedge)", "DESPESA", "DEDUCAO_INTERMEDIACAO", "", "", "Lei 9.718/98 §6 I 'e'"),
    ("8.1.7.00.00", "Despesas de Pessoal", "DESPESA", "DESPESA_OPERACIONAL", "", "", "Não reduz a base"),
    ("8.1.7.05.00", "Despesas Administrativas", "DESPESA", "DESPESA_OPERACIONAL", "", "", "Não reduz a base"),
    ("8.1.7.06.00", "Despesas Tributárias", "DESPESA", "DESPESA_OPERACIONAL", "", "", "Não reduz a base"),
    ("1.1.1.00.00", "Disponibilidades", "PATRIMONIAL", "FORA_APURACAO", "", "", "Fora da apuração"),
    ("6.1.1.00.00", "Capital Social", "PATRIMONIAL", "FORA_APURACAO", "", "", "Fora da apuração"),
]


def gerar_balancete(caminho="balancete.xlsx"):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Balancete"
    _cab(ws, ["Conta", "Descricao", "Saldo_Anterior", "Debito", "Credito", "Saldo_Atual"])
    for conta, desc, saldo in BALANCETE:
        ws.append([conta, desc, 0.0, 0.0, 0.0, saldo])
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            cel.font = Font(name="Arial", size=10)
            if cel.column >= 3:
                cel.number_format = "#,##0.00"
    _ajustar(ws); ws.freeze_panes = "A2"; wb.save(caminho); return caminho


def gerar_cadastro(caminho="cadastro.xlsx"):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Cadastro"
    _cab(ws, ["Conta", "Descricao", "Natureza", "Classificacao", "CST_PIS", "CST_COFINS", "Observacao"])
    for linha in CADASTRO:
        ws.append(list(linha))
    for row in ws.iter_rows(min_row=2):
        for cel in row:
            cel.font = Font(name="Arial", size=10)
    _ajustar(ws); ws.freeze_panes = "A2"; wb.save(caminho); return caminho


if __name__ == "__main__":
    import os, sys
    destino = sys.argv[1] if len(sys.argv) > 1 else "."
    os.makedirs(destino, exist_ok=True)
    print("Gerado:", gerar_balancete(os.path.join(destino, "balancete.xlsx")))
    print("Gerado:", gerar_cadastro(os.path.join(destino, "cadastro.xlsx")))
